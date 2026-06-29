#!/usr/bin/env python
r"""
XLSX -> CSV converter for the FRG Foerdermonitoring MVP2 interim app.

Usage examples:
  python xlsx_to_csv.py "C:\path\source.xlsx" --sheet MVP
  python xlsx_to_csv.py "C:\path\source.xlsx" --out "C:\path\output.csv"
  python xlsx_to_csv.py --list-sheets "C:\path\source.xlsx"

If started without arguments, a small Windows file dialog is opened.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable, Sequence

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "openpyxl is missing. Install with: python -m pip install openpyxl"
    ) from exc

HEADER_ALIASES = {
    "projekt id": "projekt_id",
    "projekt_id": "projekt_id",
    "project id": "projekt_id",
    "projekt": "projekt",
    "monat": "monat",
    "month": "monat",
    "ausbautreiber": "ausbautreiber",
    "kommune": "kommune",
    "foerderprogramm": "foerderprogramm",
    "förderprogramm": "foerderprogramm",
    "vertrag": "vertrag_id",
    "vertrag id": "vertrag_id",
    "vertrag_id": "vertrag_id",
    "pfau id": "vertrag_id",
    "pfau_id": "vertrag_id",
    "los": "los_id",
    "los id": "los_id",
    "los_id": "los_id",
    "hp": "hp_ist",
    "hp ist": "hp_ist",
    "hp_ist": "hp_ist",
    "bp": "bp_ist",
    "bp ist": "bp_ist",
    "bp_ist": "bp_ist",
    "hc": "hc_ist",
    "hc ist": "hc_ist",
    "hc_ist": "hc_ist",
    "adressen soll": "adressen_soll",
    "adressen_soll": "adressen_soll",
    "solladressen": "adressen_soll",
    "abrechenbar": "abrechenbar",
    "cash in ist": "cash_in_ist",
    "cash-in ist": "cash_in_ist",
    "cash_in_ist": "cash_in_ist",
    "forecast cash in": "forecast_cash_in",
    "forecast_cash_in": "forecast_cash_in",
    "forecast": "forecast_cash_in",
    "ipf cash in": "ipf_cash_in",
    "ipf_cash_in": "ipf_cash_in",
    "ipf": "ipf_cash_in",
    "ipf hc": "ipf_hc",
    "ipf_hc": "ipf_hc",
    "risiko": "risiko_hinweis",
    "risiko hinweis": "risiko_hinweis",
    "risiko_hinweis": "risiko_hinweis",
    "owner": "owner",
    "status": "status",
}


def choose_file_gui() -> Path | None:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None
    root = tk.Tk()
    root.withdraw()
    selected = filedialog.askopenfilename(
        title="Foerdermonitoring XLSX auswaehlen",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(selected) if selected else None


def clean_header(value: object, fallback_index: int) -> str:
    raw = "" if value is None else str(value).strip()
    if not raw:
        raw = f"spalte_{fallback_index}"
    lookup = re.sub(r"\s+", " ", raw.lower()).strip()
    if lookup in HEADER_ALIASES:
        return HEADER_ALIASES[lookup]
    normalized = raw.lower()
    normalized = normalized.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return normalized or f"spalte_{fallback_index}"


def make_unique(headers: Sequence[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    for header in headers:
        count = seen.get(header, 0)
        seen[header] = count + 1
        unique.append(header if count == 0 else f"{header}_{count + 1}")
    return unique


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def row_nonempty(row: Iterable[object]) -> int:
    return sum(1 for value in row if value is not None and str(value).strip() != "")


def detect_header_row(rows: list[tuple[object, ...]], max_scan: int = 40) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows[:max_scan]):
        nonempty = row_nonempty(row)
        textish = sum(1 for value in row if isinstance(value, str) and value.strip())
        score = nonempty * 2 + textish
        if score > best_score:
            best_score = score
            best_index = idx
    return best_index


def convert_xlsx(source: Path, out: Path | None, sheet_name: str | None, header_row: int | None) -> Path:
    if not source.exists():
        raise FileNotFoundError(source)
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
            ws = wb[sheet_name]
        elif "MVP" in wb.sheetnames:
            ws = wb["MVP"]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Sheet '{ws.title}' is empty")

        header_idx = (header_row - 1) if header_row else detect_header_row(rows)
        headers = make_unique([clean_header(value, i + 1) for i, value in enumerate(rows[header_idx])])
        data_rows = [row for row in rows[header_idx + 1 :] if row_nonempty(row) > 0]

        target = out or source.with_name(f"{source.stem}_{ws.title}.csv")
        target.parent.mkdir(parents=True, exist_ok=True)
        with target.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, delimiter=";")
            writer.writerow(headers)
            for row in data_rows:
                padded = list(row) + [None] * (len(headers) - len(row))
                writer.writerow([cell_to_text(value) for value in padded[: len(headers)]])
        return target
    finally:
        wb.close()


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Convert XLSX/XLSM sheets to semicolon CSV for the MVP2 interim app.")
    parser.add_argument("xlsx", nargs="?", help="Path to .xlsx/.xlsm file. If omitted, a file dialog opens.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to 'MVP' if present, otherwise first sheet.")
    parser.add_argument("--out", help="Output CSV path. Defaults to <xlsx>_<sheet>.csv")
    parser.add_argument("--header-row", type=int, help="1-based header row. Defaults to auto-detection.")
    parser.add_argument("--list-sheets", action="store_true", help="List workbook sheets and exit.")
    args = parser.parse_args(argv)

    source = Path(args.xlsx) if args.xlsx else choose_file_gui()
    if source is None:
        print("No file selected.", file=sys.stderr)
        return 2

    if args.list_sheets:
        wb = load_workbook(source, read_only=True, data_only=True)
        try:
            for name in wb.sheetnames:
                print(name)
        finally:
            wb.close()
        return 0

    target = convert_xlsx(source, Path(args.out) if args.out else None, args.sheet, args.header_row)
    print(f"CSV written: {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


